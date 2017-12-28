from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

W_1 = 0.2
W_2 = 0.4
W_3 = 0.4

wb = load_workbook('Source - Python.xlsx')

# Address a named worksheet
ws_telephone = wb['Telephone']
ws_mobile = wb['Mobile']
ws_internet = wb['Internet']
# Read value of B4

data = {}

#data['Afghanistan'] = {2006 : {'telephone' : 2008, 'internet' : 45}}
#data['Afghanistan'][2007] = {'apple': 23}

for x in range(3, 191):
    for y in range(2, 18):
        country = ws_internet['A' + str(x)].value
        year = ws_internet[get_column_letter(y) + '2'].value
        data[country] = {year: 0}

for x in range(3, 191):
    for y in range(2, 18):
        cell_internet = ws_internet[get_column_letter(y) + str(x)].value
        if cell_internet is None:
            cell_internet = 0
        cell_mobile = ws_mobile[get_column_letter(y) + str(x)].value
        if cell_mobile is None:
            cell_mobile = 0
        cell_telephone = ws_telephone[get_column_letter(y) + str(x)].value
        if cell_telephone is None:
            cell_telephone = 0
        country = ws_internet['A' + str(x)].value
        year = ws_internet[get_column_letter(y) + '2'].value
        data[country][year] = {'telephone': cell_telephone,
                               'mobile': cell_mobile,
                               'internet': cell_internet}


print(data)
print(data['Germany'][2006]['telephone'])
print(max(data[n][2006]['mobile'] for n in data))


def calculate_idi(data, country, year):
    f_1 = data[country][year]['telephone']
    f_2 = data[country][year]['mobile']
    f_3 = data[country][year]['internet']
    max_f1 = max(data[n][year]['telephone'] for n in data)
    max_f2 = max(data[n][year]['mobile'] for n in data)
    max_f3 = max(data[n][year]['internet'] for n in data)
    idi = ((f_1/max_f1)*W_1) + ((f_2/max_f2)*W_2) + ((f_3/max_f3)*W_3)
    return idi


def update_data_idi(data):
    for x in data:
        for y in data[x]:
            data[x][y]['idi'] = calculate_idi(data,x,y)

    return data


print(calculate_idi(data, 'Germany', 2006))
data = update_data_idi(data)
print(data)







cell = ws_internet[get_column_letter(2) + "4"].value
# If empty cell in Excel, the value will be None
if not cell is None:
    print(cell)


